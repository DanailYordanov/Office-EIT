import os
import shutil
import secrets
from django.conf import settings
from django.views.generic import DeleteView
from django.urls import reverse, reverse_lazy
from django.contrib.auth import get_user_model
from django.http import JsonResponse, HttpResponse
from django.core.exceptions import PermissionDenied
from django.utils import timezone, dateformat, formats
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from pyVies import api
from openpyxl import load_workbook
from num2cyrillic import NumberToWords
from deep_translator import GoogleTranslator
from django_select2.views import AutoResponseView
from main import models
from main import forms


@login_required
def cars_list(request):
    if request.user.is_staff:
        cars = models.Car.objects.all()
    else:
        raise PermissionDenied

    context = {
        'cars': cars,
        'page_heading': 'Автомобили'
    }

    return render(request, 'main/lists/cars_list.html', context)


@login_required
def add_car(request):
    if request.user.is_staff:
        if request.method == 'POST':
            form = forms.CarModelForm(request.POST)

            if form.is_valid():
                form.save()
                return redirect('main:cars-list')
        else:
            form = forms.CarModelForm()
    else:
        raise PermissionDenied

    context = {
        'form': form,
        'url': reverse('main:add-car'),
        'page_heading': 'Добавяне на автомобил'
    }

    return render(request, 'main/add_update_form.html', context)


@login_required
def update_car(request, pk):
    if request.user.is_staff:
        car = get_object_or_404(models.Car, id=pk)

        if request.method == 'POST':
            form = forms.CarModelForm(request.POST, instance=car)

            if form.is_valid():
                form.save()
                return redirect('main:cars-list')
        else:
            form = forms.CarModelForm(instance=car)
    else:
        raise PermissionDenied

    context = {
        'form': form,
        'url': reverse('main:update-car', args=(pk,)),
        'page_heading': 'Редактиране на автомобил'
    }

    return render(request, 'main/add_update_form.html', context)


class CarDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):

    model = models.Car
    success_url = reverse_lazy('main:cars-list')

    def test_func(self):
        return self.request.user.is_staff


@login_required
def reminders_list(request):
    if request.user.is_staff:
        reminders = models.Reminder.objects.all()
    else:
        raise PermissionDenied

    context = {
        'reminders': reminders,
        'page_heading': 'Напомняния'
    }

    return render(request, 'main/lists/reminders_list.html', context)


@login_required
def add_reminder(request):
    if request.user.is_staff:
        if request.method == 'POST':
            form = forms.ReminderModelForm(request.POST)

            if form.is_valid():
                form.save()
                return redirect('main:reminders-list')
        else:
            form = forms.ReminderModelForm()
    else:
        raise PermissionDenied

    context = {
        'form': form,
        'url': reverse('main:add-reminder'),
        'page_heading': 'Добавяне на напомняне'
    }

    return render(request, 'main/add_update_form.html', context)


@login_required
def update_reminder(request, pk):
    if request.user.is_staff:
        reminder = get_object_or_404(models.Reminder, id=pk)

        if request.method == 'POST':
            form = forms.ReminderModelForm(request.POST, instance=reminder)

            if form.is_valid():
                form.save()
                return redirect('main:reminders-list')
        else:
            form = forms.ReminderModelForm(instance=reminder)
    else:
        raise PermissionDenied

    context = {
        'form': form,
        'url': reverse('main:update-reminder', args=(pk,)),
        'page_heading': 'Редактиране на напомняне'
    }

    return render(request, 'main/add_update_form.html', context)


class ReminderDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):

    model = models.Reminder
    success_url = reverse_lazy('main:reminders-list')

    def test_func(self):
        return self.request.user.is_staff


@login_required
def services_list(request):
    if request.user.is_staff:
        services = models.Service.objects.all()
    else:
        raise PermissionDenied

    context = {
        'services': services,
        'page_heading': 'Технически обслужвания'
    }

    return render(request, 'main/lists/services_list.html', context)


@login_required
def add_service(request):
    if request.user.is_staff:
        if request.method == 'POST':
            form = forms.ServiceModelForm(request.POST)

            if form.is_valid():
                form.save()
                return redirect('main:services-list')
        else:
            form = forms.ServiceModelForm()
    else:
        raise PermissionDenied

    context = {
        'form': form,
        'url': reverse('main:add-service'),
        'page_heading': 'Добавяне на техническо обслужване'
    }

    return render(request, 'main/add_update_form.html', context)


@login_required
def update_service(request, pk):
    if request.user.is_staff:
        service = get_object_or_404(models.Service, id=pk)

        if request.method == 'POST':
            form = forms.ServiceModelForm(request.POST, instance=service)

            if form.is_valid():
                form.save()
                return redirect('main:services-list')
        else:
            form = forms.ServiceModelForm(instance=service)
    else:
        raise PermissionDenied

    context = {
        'form': form,
        'url': reverse('main:update-service', args=(pk,)),
        'page_heading': 'Редактиране на техническо обслужване'
    }

    return render(request, 'main/add_update_form.html', context)


class ServiceDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):

    model = models.Service
    success_url = reverse_lazy('main:services-list')

    def test_func(self):
        return self.request.user.is_staff


@login_required
def users_details_list(request):
    if request.user.is_staff:
        users = get_user_model().objects.filter(is_staff=False, is_active=True)
    else:
        raise PermissionDenied

    context = {
        'users': users,
        'page_heading': 'Документи на шофьори'
    }

    return render(request, 'main/lists/users_details_list.html', context)


@login_required
def user_set_unactive(request, pk=None):
    if request.user.is_staff:
        if request.method == 'POST':
            user = get_object_or_404(get_user_model(), pk=pk)
            user.is_active = False
            user.save()

        return redirect('main:users-details-list')
    else:
        raise PermissionDenied


@login_required
def contractors_list(request):
    if request.user.is_staff:
        contractors = models.Contractor.objects.all()
    else:
        raise PermissionDenied

    context = {
        'contractors': contractors,
        'page_heading': 'Контрагенти'
    }

    return render(request, 'main/lists/contractors_list.html', context)


@login_required
def add_contractor(request):
    if request.user.is_staff:
        if request.method == 'POST':
            form = forms.ContractorsModelForm(request.POST, request.FILES)

            if form.is_valid():
                form.save()
                return redirect('main:contractors-list')
        else:
            form = forms.ContractorsModelForm()
    else:
        raise PermissionDenied

    context = {
        'form': form,
        'url': reverse('main:add-contractor'),
        'page_heading': 'Добавяне на контрагент'
    }

    return render(request, 'main/add_update_form.html', context)


@login_required
def update_contractor(request, pk):
    if request.user.is_staff:
        contractor = get_object_or_404(models.Contractor, id=pk)

        if request.method == 'POST':
            form = forms.ContractorsModelForm(
                request.POST, request.FILES, instance=contractor)

            if form.is_valid():
                form.save()
                return redirect('main:contractors-list')
        else:
            form = forms.ContractorsModelForm(instance=contractor)
    else:
        raise PermissionDenied

    context = {
        'form': form,
        'url': reverse('main:update-contractor', args=(pk,)),
        'page_heading': 'Редактиране на контрагент'
    }

    return render(request, 'main/add_update_form.html', context)


class ContractorDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):

    model = models.Contractor
    success_url = reverse_lazy('main:contractors-list')

    def test_func(self):
        return self.request.user.is_staff


@login_required
def populate_vat_info(request):
    if request.method == 'POST':
        bulstat = request.POST.get('bulstat')

        data = {
            'name': None,
            'city': None,
            'address': None,
            'postal_code': None
        }

        if bulstat != '':
            try:
                vies = api.Vies()
                result = vies.request(bulstat, extended_info=True)

                data['name'] = result['traderName']
                data['city'] = result['traderCity']
                data['address'] = result['traderAddress']
                data['postal_code'] = result['traderPostcode']
            except Exception:
                pass

        return JsonResponse(data)
    else:
        raise PermissionDenied


@login_required
def courses_list(request):
    if request.user.is_staff:
        courses = models.Course.objects.all()
    else:
        courses = models.Course.objects.filter(
            driver=request.user)

    context = {
        'courses': courses,
        'page_heading': 'Курсове'
    }

    return render(request, 'main/lists/courses_list.html', context)


@login_required
def add_course(request, pk=None):
    if request.user.is_staff:

        initial_form = dict()
        initial_formset = list()

        if pk:
            initial_course = get_object_or_404(models.Course, id=pk)

            initial_form['driver'] = initial_course.driver
            initial_form['car'] = initial_course.car
            initial_form['company'] = initial_course.company
            initial_form['contractor'] = initial_course.contractor
            initial_form['bank'] = initial_course.bank
            initial_form['from_to'] = initial_course.from_to
            initial_form['cargo_type'] = initial_course.cargo_type
            initial_form['export'] = initial_course.export
            initial_form['contact_person'] = initial_course.contact_person
            initial_form['other_conditions'] = initial_course.other_conditions

            if hasattr(initial_course, 'medical_examination'):
                initial_form['medical_examination_perpetrator'] = initial_course.medical_examination

            if hasattr(initial_course, 'technical_inspection'):
                initial_form['technical_inspection_perpetrator'] = initial_course.technical_inspection

            for address in initial_course.addresses.all():
                initial_formset.append(
                    {
                        'load_type': address.load_type,
                        'address': address.address
                    }
                )

            forms.CourseAddressAddFormset.extra = len(initial_formset)

        else:
            medical_examination_perpetrator = models.MedicalExaminationPerpetrator.objects.all().last()
            initial_form['medical_examination_perpetrator'] = medical_examination_perpetrator

            technical_inspection_perpetrator = models.TechnicalInspectionPerpetrator.objects.all().last()
            initial_form['technical_inspection_perpetrator'] = technical_inspection_perpetrator

        if request.method == 'POST':
            form = forms.CourseModelForm(request.POST, initial=initial_form)
            formset = forms.CourseAddressAddFormset(
                request.POST, initial=initial_formset)

            if form.is_valid() and formset.is_valid():
                form_instance = form.save()

                for f in formset:
                    if f.is_valid():
                        f.instance.course = form_instance
                        f.save()

                return redirect('main:courses-list')
        else:
            form = forms.CourseModelForm(initial=initial_form)
            formset = forms.CourseAddressAddFormset(initial=initial_formset)
    else:
        raise PermissionDenied

    context = {
        'form': form,
        'formset': formset,
        'url': reverse('main:add-course'),
        'page_heading': 'Добавяне на курс'
    }

    return render(request, 'main/add_update_course.html', context)


@login_required
def update_course(request, pk):
    if request.user.is_staff:
        course = get_object_or_404(models.Course, id=pk)

        if request.method == 'POST':
            form = forms.CourseModelForm(request.POST, instance=course)
            formset = forms.CourseAddressUpdateFormset(
                request.POST, instance=course)

            if form.is_valid() and formset.is_valid():
                form_instance = form.save()

                for f in formset:
                    if f.is_valid():
                        f.instance.course = form_instance
                        f.save()

                formset.save()
                return redirect('main:courses-list')
        else:
            form = forms.CourseModelForm(instance=course)
            formset = forms.CourseAddressUpdateFormset(instance=course)

        context = {
            'form': form,
            'formset': formset,
            'url': reverse('main:update-course', args=(pk,)),
            'page_heading': 'Редактиране на курс'
        }

        return render(request, 'main/add_update_course.html', context)
    else:
        raise PermissionDenied


@login_required
def course_information(request, pk):
    course = get_object_or_404(models.Course, id=pk)
    expenses = models.Expense.objects.filter(course=course)
    addresses = models.CourseAddress.objects.filter(course=course)

    context = {
        'course': course,
        'expenses': expenses,
        'addresses': addresses,
        'page_heading': 'Информация за курс'
    }

    return render(request, 'main/course_information.html', context)


class CourseDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):

    model = models.Course
    success_url = reverse_lazy('main:courses-list')

    def test_func(self):
        return self.request.user.is_staff


def load_contractor_reminder(request):
    if request.method == 'POST':
        contractor_id = request.POST.get('contractor_id')

        data = {
            'cmr_reminder': None,
            'license_reminder': None
        }

        if contractor_id != '':
            contractor = get_object_or_404(
                models.Contractor, id=int(contractor_id))

            if contractor.cmr_expiration_date:
                if contractor.cmr_expiration_date <= (timezone.now().date() + timezone.timedelta(days=5)):
                    data['cmr_reminder'] = f'ЧМР Застраховката изтича на {contractor.cmr_expiration_date}'

            if contractor.license_expiration_date:
                if contractor.license_expiration_date <= (timezone.now().date() + timezone.timedelta(days=5)):
                    data['license_reminder'] = f'Лиценза изтича на {contractor.license_expiration_date}'

        return JsonResponse(data)
    else:
        raise PermissionDenied


@login_required
def addresses_list(request):
    if request.user.is_staff:
        addresses = models.Address.objects.all()
    else:
        raise PermissionDenied

    context = {
        'addresses': addresses,
        'page_heading': 'Адреси'
    }

    return render(request, 'main/lists/addresses_list.html', context)


@login_required
def add_address(request):
    if request.user.is_staff:
        if request.method == 'POST':
            form = forms.AddressModelForm(request.POST)

            if form.is_valid():
                form.save()
                return redirect('main:addresses-list')
        else:
            form = forms.AddressModelForm()
    else:
        raise PermissionDenied

    context = {
        'form': form,
        'url': reverse('main:add-address'),
        'page_heading': 'Добавяне на адрес'
    }

    return render(request, 'main/add_update_form.html', context)


@login_required
def update_address(request, pk):
    if request.user.is_staff:
        address = get_object_or_404(models.Address, id=pk)

        if request.method == 'POST':
            form = forms.AddressModelForm(request.POST, instance=address)

            if form.is_valid():
                form.save()
                return redirect('main:addresses-list')
        else:
            form = forms.AddressModelForm(instance=address)
    else:
        raise PermissionDenied

    context = {
        'form': form,
        'url': reverse('main:update-address', args=(pk,)),
        'page_heading': 'Редактиране на адрес'
    }

    return render(request, 'main/add_update_form.html', context)


class AddressDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):

    model = models.Address
    success_url = reverse_lazy('main:addresses-list')

    def test_func(self):
        return self.request.user.is_staff


@login_required
def add_expense(request, pk):
    course = get_object_or_404(models.Course, id=pk)

    if request.method == 'POST':
        form = forms.ExpenseModelForm(request.POST)

        if form.is_valid():
            form.instance.course = course
            form.save()
            return redirect('main:course-information', pk=pk)
    else:
        form = forms.ExpenseModelForm()

    context = {
        'form': form,
        'url': reverse('main:add-expense', args=(pk,)),
        'page_heading': 'Добавяне на разход'
    }

    return render(request, 'main/add_update_form.html', context)


@login_required
def update_expense(request, pk):
    expense = get_object_or_404(models.Expense, id=pk)

    if request.method == 'POST':
        form = forms.ExpenseModelForm(request.POST, instance=expense)

        if form.is_valid():
            form.save()
            return redirect('main:course-information', pk=expense.course.id)
    else:
        form = forms.ExpenseModelForm(instance=expense)

    context = {
        'form': form,
        'url': reverse('main:update-expense', args=(pk,)),
        'page_heading': 'Редактиране на разход'
    }

    return render(request, 'main/add_update_form.html', context)


class ExpenseDeleteView(LoginRequiredMixin, DeleteView):

    model = models.Expense

    def get_success_url(self):
        return reverse_lazy('main:course-information', kwargs={'pk': self.kwargs['course_pk']})


@login_required
def trip_orders_list(request):
    if request.user.is_staff:
        trip_orders = models.TripOrder.objects.all()
    else:
        raise PermissionDenied

    context = {
        'trip_orders': trip_orders,
        'page_heading': 'Командировъчни заповеди'
    }

    return render(request, 'main/lists/trip_orders_list.html', context)


@login_required
def add_trip_order(request):
    if request.user.is_staff:
        if request.method == 'POST':
            form = forms.TripOrderModelForm(request.POST)

            if form.is_valid():
                form.instance.creator = request.user
                form.save()

                return redirect('main:trip-orders-list')
        else:
            form = forms.TripOrderModelForm()
    else:
        raise PermissionDenied

    context = {
        'form': form,
        'url': reverse('main:add-trip-order'),
        'page_heading': 'Добавяне на командировъчна заповед'
    }

    return render(request, 'main/add_update_form.html', context)


@login_required
def update_trip_order(request, pk):
    trip_order = get_object_or_404(models.TripOrder, id=pk)

    if request.method == 'POST':
        form = forms.TripOrderModelForm(request.POST, instance=trip_order)

        if form.is_valid():
            form.save()
            return redirect('main:trip-orders-list')
    else:
        form = forms.TripOrderModelForm(instance=trip_order)

    context = {
        'form': form,
        'url': reverse('main:update-trip-order', args=(pk,)),
        'page_heading': 'Редактиране на командировъчна заповед'
    }

    return render(request, 'main/add_update_form.html', context)


class TripOrderDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):

    model = models.TripOrder
    success_url = reverse_lazy('main:trip-orders-list')

    def test_func(self):
        return self.request.user.is_staff


@login_required
def trip_order_xlsx(request, pk):
    if request.user.is_staff:
        trip_order = get_object_or_404(models.TripOrder, id=pk)

        unique_token = secrets.token_hex(32)

        unique_dir_path = os.path.join(
            settings.BASE_DIR, f'main/xlsx_files/{unique_token}')

        os.mkdir(unique_dir_path)

        trip_order_xlsx_path = os.path.join(
            settings.BASE_DIR, 'main/xlsx_files/trip_order.xlsx')

        unique_trip_order_xlsx_path = os.path.join(
            unique_dir_path, 'trip_order.xlsx')

        shutil.copy(trip_order_xlsx_path, unique_trip_order_xlsx_path)

        duration_time = (trip_order.to_date - trip_order.from_date).days + 1
        company = trip_order.course.company
        heading = f'{company.name}, {company.city}, ЕИК {company.bulstat}'

        wb = load_workbook(filename=unique_trip_order_xlsx_path)
        ws = wb.active

        ws['A1'] = heading
        ws['E3'] = trip_order.number
        ws['G3'] = dateformat.format(
            trip_order.creation_date, formats.get_format('SHORT_DATE_FORMAT'))
        ws['A8'] = trip_order.driver.__str__()
        ws['B12'] = trip_order.destination
        ws['C13'] = dateformat.format(
            trip_order.from_date, formats.get_format('SHORT_DATE_FORMAT'))
        ws['E13'] = dateformat.format(
            trip_order.to_date, formats.get_format('SHORT_DATE_FORMAT'))
        ws['I13'] = duration_time
        ws['F15'] = trip_order.course.car.number_plate
        ws['F18'] = f'{company.name}'
        ws['A29'] = trip_order.driver.__str__()
        ws['F29'] = trip_order.creator.__str__()

        wb.save(unique_trip_order_xlsx_path)

        course = trip_order.course

        course_expenses_xlsx_path = os.path.join(
            settings.BASE_DIR, 'main/xlsx_files/course_expenses.xlsx')

        unique_course_expenses_xlsx_path = os.path.join(
            unique_dir_path, 'course_expenses.xlsx')

        shutil.copy(course_expenses_xlsx_path,
                    unique_course_expenses_xlsx_path)

        wb = load_workbook(filename=unique_course_expenses_xlsx_path)
        ws = wb.active

        ws['B1'] = trip_order.number
        ws['E1'] = trip_order.from_date
        ws['B2'] = course.car.number_plate
        ws['G2'] = course.driver.__str__()
        ws['G17'] = course.driver.debit_card_number

        wb.save(unique_course_expenses_xlsx_path)

        shutil.make_archive(unique_dir_path, 'zip', unique_dir_path)

        zip_path = os.path.join(
            settings.BASE_DIR, f'main/xlsx_files/{unique_token}.zip')

        response = HttpResponse(open(zip_path, 'rb'))
        response['Content-Type'] = 'application/zip'
        response['Content-Disposition'] = f'attachment; filename="Trip Order {trip_order.number}.zip"'

        os.remove(zip_path)
        shutil.rmtree(unique_dir_path, ignore_errors=True)

        return response
    else:
        raise PermissionDenied


@login_required
def load_dates(request):
    if request.method == 'POST':
        course_id = request.POST.get('course_id')

        data = {
            'from_date': None,
            'to_date': None
        }

        if course_id != '':
            course_addresses = models.CourseAddress.objects.filter(
                course__id=int(course_id))

            from_date = course_addresses.filter(
                load_type='Адрес на товарене').order_by('date')
            to_date = course_addresses.filter(
                load_type='Адрес на разтоварване').order_by('date')

            if (from_date):
                data['from_date'] = from_date.first().date

            if (to_date):
                data['to_date'] = to_date.last().date

        return JsonResponse(data)
    else:
        raise PermissionDenied


@login_required
def expense_orders_list(request):
    if request.user.is_staff:
        expense_orders = models.ExpenseOrder.objects.all()
    else:
        raise PermissionDenied

    context = {
        'expense_orders': expense_orders,
        'page_heading': 'Разходни ордери'
    }

    return render(request, 'main/lists/expense_orders_list.html', context)


@login_required
def add_expense_order(request):
    if request.user.is_staff:
        if request.method == 'POST':
            form = forms.ExpenseOrderModelForm(request.POST)

            if form.is_valid():
                form.instance.creator = request.user
                form.save()

                return redirect('main:expense-orders-list')
        else:
            form = forms.ExpenseOrderModelForm()

        context = {
            'form': form,
            'url': reverse('main:add-expense-order'),
            'page_heading': 'Добавяне на разходен ордер'
        }

        return render(request, 'main/add_update_form.html', context)
    else:
        raise PermissionDenied


@login_required
def update_expense_order(request, pk):
    if request.user.is_staff:
        expense_order = get_object_or_404(models.ExpenseOrder, id=pk)

        if request.method == 'POST':
            form = forms.ExpenseOrderModelForm(
                request.POST, instance=expense_order)

            if form.is_valid():
                form.save()
                return redirect('main:expense-orders-list')
        else:
            form = forms.ExpenseOrderModelForm(instance=expense_order)

        context = {
            'form': form,
            'url': reverse('main:update-expense-order', args=(pk,)),
            'page_heading': 'Редактиране на разходен ордер'
        }

        return render(request, 'main/add_update_form.html', context)
    else:
        raise PermissionDenied


class ExpenseOrderDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):

    model = models.ExpenseOrder
    success_url = reverse_lazy('main:expense-orders-list')

    def test_func(self):
        return self.request.user.is_staff


@login_required
def expense_order_xlsx(request, pk):
    if request.user.is_staff:
        expense_order = get_object_or_404(models.ExpenseOrder, id=pk)

        unique_token = secrets.token_hex(32)

        xlsx_path = os.path.join(
            settings.BASE_DIR, 'main/xlsx_files/expense_order.xlsx')

        unique_xlsx_path = os.path.join(
            settings.BASE_DIR, f'main/xlsx_files/expense_order_{unique_token}.xlsx')

        shutil.copy(xlsx_path, unique_xlsx_path)

        company = expense_order.trip_order.course.company
        heading = f'{company.name}, {company.city}, ЕИК {company.bulstat}'

        wb = load_workbook(filename=unique_xlsx_path)
        ws = wb.active

        ws['A1'] = heading
        ws['E3'] = expense_order.number
        ws['G3'] = dateformat.format(
            expense_order.creation_date, formats.get_format('SHORT_DATE_FORMAT'))
        ws['D5'] = expense_order.trip_order.driver.__str__()
        ws['B7'] = expense_order.BGN_amount
        ws['E7'] = expense_order.EUR_amount
        ws['F10'] = expense_order.trip_order.number
        ws['H10'] = dateformat.format(
            expense_order.trip_order.from_date, formats.get_format('SHORT_DATE_FORMAT'))
        ws['G12'] = expense_order.trip_order.driver.debit_card_number

        if expense_order.trip_order.driver.bank:
            ws['E13'] = expense_order.trip_order.driver.bank.iban

        ws['B15'] = dateformat.format(
            expense_order.creation_date, formats.get_format('SHORT_DATE_FORMAT'))
        ws['A19'] = expense_order.trip_order.driver.__str__()
        ws['F19'] = expense_order.creator.__str__()

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="Expense Order {expense_order.number}.xlsx"'

        wb.save(response)

        os.remove(unique_xlsx_path)

        return response
    else:
        raise PermissionDenied


@login_required
def course_invoices_list(request):
    if request.user.is_staff:
        course_invoices = models.CourseInvoice.objects.all()
    else:
        raise PermissionDenied

    context = {
        'course_invoices': course_invoices,
        'page_heading': 'Фактури за курсове'
    }

    return render(request, 'main/lists/course_invoices_list.html', context)


@login_required
def add_course_invoice(request):
    if request.user.is_staff:
        if request.method == 'POST':
            form = forms.CourseInvoiceModelForm(request.POST)

            if form.is_valid():
                form.instance.creator = request.user
                form.save()

                return redirect('main:course-invoices-list')
        else:
            form = forms.CourseInvoiceModelForm()
    else:
        raise PermissionDenied

    context = {
        'form': form,
        'url': reverse('main:add-course-invoice'),
        'page_heading': 'Добавяне на фактура за курс'
    }

    return render(request, 'main/add_update_form.html', context)


@login_required
def update_course_invoice(request, pk):
    course_invoice = get_object_or_404(models.CourseInvoice, id=pk)

    if request.method == 'POST':
        form = forms.CourseInvoiceModelForm(
            request.POST, instance=course_invoice)

        if form.is_valid():
            form.save()
            return redirect('main:course-invoices-list')
    else:
        form = forms.CourseInvoiceModelForm(instance=course_invoice)

    context = {
        'form': form,
        'url': reverse('main:update-course-invoice', args=(pk,)),
        'page_heading': 'Редактиране на фактура за курс'
    }

    return render(request, 'main/add_update_form.html', context)


class CourseInvoiceDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):

    model = models.CourseInvoice
    success_url = reverse_lazy('main:course-invoices-list')

    def test_func(self):
        return self.request.user.is_staff


@login_required
def course_invoice_xlsx(request, pk):
    if request.user.is_staff:
        course_invoice = get_object_or_404(models.CourseInvoice, id=pk)

        unique_token = secrets.token_hex(32)

        unique_dir_path = os.path.join(
            settings.BASE_DIR, f'main/xlsx_files/{unique_token}')

        os.mkdir(unique_dir_path)

        original_xlsx_path = os.path.join(
            settings.BASE_DIR, 'main/xlsx_files/course_invoice.xlsx')

        original_unique_xlsx_path = os.path.join(
            unique_dir_path, 'course_invoice_original.xlsx')

        copy_unique_xlsx_path = os.path.join(
            unique_dir_path, 'course_invoice_copy.xlsx')

        shutil.copy(original_xlsx_path, original_unique_xlsx_path)

        course = course_invoice.course
        contractor = course.contractor
        bank = course.bank
        company = course.company

        currency = course.course_price_currency
        price = round(course.course_price, 2)
        vat_price = round(course.course_price * 0.2, 2)

        if course_invoice.tax_type == 'Стандартна фактура':
            calculated_price = price + vat_price
        else:
            calculated_price = price

        whole_part = int(calculated_price)
        fractional_part = int(((calculated_price - whole_part) * 100))
        price_in_words = NumberToWords()

        price = f'{price} {currency}'
        vat_price = f'{vat_price} {currency}'
        calculated_price = f'{calculated_price} {currency}'

        wb = load_workbook(filename=original_unique_xlsx_path)
        ws = wb.active

        ws['V5'] = str(course_invoice.number).zfill(10)
        ws['V6'] = dateformat.format(
            course_invoice.creation_date, formats.get_format('SHORT_DATE_FORMAT'))

        ws['G9'] = contractor.name
        ws['G11'] = contractor.bulstat

        if contractor.bulstat:
            if contractor.client_type == 'Български':
                ws['G12'] = contractor.bulstat[2:]
            else:
                ws['G12'] = contractor.bulstat

        ws['G13'] = contractor.city
        ws['G14'] = contractor.address
        ws['G16'] = contractor.mol
        ws['G17'] = contractor.phone_number

        ws['R9'] = company.name

        if company.bulstat:
            ws['R11'] = company.bulstat
            ws['R12'] = company.bulstat[2:]

        ws['R13'] = company.city
        ws['R14'] = company.address
        ws['R16'] = company.mol
        ws['R17'] = company.phone_number

        ws['T21'] = price

        if course_invoice.tax_type == 'Стандартна фактура':
            ws['X24'] = price
            ws['X25'] = vat_price

        ws['X26'] = calculated_price

        if fractional_part != 0:
            ws['G24'] = f'{price_in_words.cyrillic(whole_part)} лева и {price_in_words.cyrillic(fractional_part)} стотинки'
        else:
            ws['G24'] = f'{price_in_words.cyrillic(whole_part)} лева'

        from_to = f'Транспорт от {course.from_to.__str__()} с камион'

        ws['H27'] = course_invoice.additional_information

        ws['C21'] = from_to
        ws['I22'] = course.car.number_plate
        ws['H23'] = course.request_number.__str__()

        ws['J28'] = course_invoice.creation_date
        ws['J29'] = course_invoice.tax_transaction_basis.__str__()

        ws['R28'] = course_invoice.payment_type
        ws['R29'] = bank.iban
        ws['R31'] = bank.name
        ws['R33'] = bank.bank_code

        ws['I35'] = course.contact_person
        ws['S35'] = course_invoice.creator.__str__()

        wb.save(original_unique_xlsx_path)

        shutil.copy(original_unique_xlsx_path, copy_unique_xlsx_path)

        wb = load_workbook(filename=copy_unique_xlsx_path)
        ws = wb.active

        ws['L3'] = 'Копие'

        wb.save(copy_unique_xlsx_path)

        if course.export:

            translated_invoice_xlsx_path = os.path.join(
                settings.BASE_DIR, 'main/xlsx_files/course_translated_invoice.xlsx')

            unique_translated_invoice_xlsx_path = os.path.join(
                unique_dir_path, 'course_translated_invoice.xlsx')

            shutil.copy(translated_invoice_xlsx_path,
                        unique_translated_invoice_xlsx_path)

            wb = load_workbook(filename=unique_translated_invoice_xlsx_path)
            ws = wb.active

            translator = GoogleTranslator(source='bg', target='en')

            ws['B2'] = translator.translate(company.name)
            ws['Q2'] = str(course_invoice.number).zfill(10)
            ws['C3'] = company.bulstat

            if company.address:
                ws['B4'] = translator.translate(company.address)

            if company.city:    
                ws['B5'] = translator.translate(company.city)

            ws['C17'] = contractor.name
            ws['C18'] = contractor.address
            ws['C19'] = contractor.bulstat

            ws['O17'] = dateformat.format(
                course_invoice.creation_date, formats.get_format('SHORT_DATE_FORMAT'))
            ws['O18'] = course.car.number_plate

            ws['D41'] = translator.translate(course.from_to.__str__())
            ws['E42'] = course.car.number_plate
            ws['N40'] = price

            if course_invoice.tax_type == 'Стандартна фактура':
                ws['P52'] = vat_price

            ws['P53'] = calculated_price

            ws['D73'] = dateformat.format(
                course_invoice.creation_date, formats.get_format('SHORT_DATE_FORMAT'))

            ws['C82'] = translator.translate(course_invoice.payment_type)
            ws['B83'] = translator.translate(bank.name)
            ws['C84'] = bank.iban
            ws['C85'] = bank.bank_code

            wb.save(unique_translated_invoice_xlsx_path)

        shutil.make_archive(unique_dir_path, 'zip', unique_dir_path)

        zip_path = os.path.join(
            settings.BASE_DIR, f'main/xlsx_files/{unique_token}.zip')

        response = HttpResponse(open(zip_path, 'rb'))
        response['Content-Type'] = 'application/zip'
        response['Content-Disposition'] = f'attachment; filename="Course Invoice {course_invoice.number}.zip"'

        os.remove(zip_path)
        shutil.rmtree(unique_dir_path, ignore_errors=True)

        return response
    else:
        raise PermissionDenied


@login_required
def companies_list(request):
    if request.user.is_staff:
        companies = models.Company.objects.all()
    else:
        raise PermissionDenied

    context = {
        'companies': companies,
        'page_heading': 'Фирми'
    }

    return render(request, 'main/lists/companies_list.html', context)


@login_required
def add_company(request):
    if request.user.is_staff:
        if request.method == 'POST':
            form = forms.CompanyModelForm(request.POST)

            if form.is_valid():
                form.save()
                return redirect('main:companies-list')
        else:
            form = forms.CompanyModelForm()
    else:
        raise PermissionDenied

    context = {
        'form': form,
        'url': reverse('main:add-company'),
        'page_heading': 'Добавяне на фирма'
    }

    return render(request, 'main/add_update_form.html', context)


@login_required
def update_company(request, pk):
    company = get_object_or_404(models.Company, id=pk)

    if request.method == 'POST':
        form = forms.CompanyModelForm(
            request.POST, instance=company)

        if form.is_valid():
            form.save()
            return redirect('main:companies-list')
    else:
        form = forms.CompanyModelForm(instance=company)

    context = {
        'form': form,
        'url': reverse('main:update-company', args=(pk,)),
        'page_heading': 'Редактиране на фирма'
    }

    return render(request, 'main/add_update_form.html', context)


class CompanyDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):

    model = models.Company
    success_url = reverse_lazy('main:companies-list')

    def test_func(self):
        return self.request.user.is_staff


@login_required
def banks_list(request):
    if request.user.is_staff:
        banks = models.Bank.objects.all()
    else:
        raise PermissionDenied

    context = {
        'banks': banks,
        'page_heading': 'Банки'
    }

    return render(request, 'main/lists/banks_list.html', context)


@login_required
def add_bank(request):
    if request.user.is_staff:
        if request.method == 'POST':
            form = forms.BankModelForm(request.POST)

            if form.is_valid():
                form.save()
                return redirect('main:banks-list')
        else:
            form = forms.BankModelForm()
    else:
        raise PermissionDenied

    context = {
        'form': form,
        'url': reverse('main:add-bank'),
        'page_heading': 'Добавяне на банка'
    }

    return render(request, 'main/add_update_form.html', context)


@login_required
def update_bank(request, pk):
    bank = get_object_or_404(models.Bank, id=pk)

    if request.method == 'POST':
        form = forms.BankModelForm(
            request.POST, instance=bank)

        if form.is_valid():
            form.save()
            return redirect('main:banks-list')
    else:
        form = forms.BankModelForm(instance=bank)

    context = {
        'form': form,
        'url': reverse('main:update-bank', args=(pk,)),
        'page_heading': 'Редактиране на банка'
    }

    return render(request, 'main/add_update_form.html', context)


class BankDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):

    model = models.Bank
    success_url = reverse_lazy('main:banks-list')

    def test_func(self):
        return self.request.user.is_staff


@login_required
def instructions_list(request):
    if request.user.is_staff:
        instructions = models.Instruction.objects.all()
    else:
        raise PermissionDenied

    context = {
        'instructions': instructions,
        'page_heading': 'Инструкции'
    }

    return render(request, 'main/lists/instructions_list.html', context)


@login_required
def add_instruction(request):
    if request.user.is_staff:
        if request.method == 'POST':
            form = forms.InstructionModelForm(request.POST)

            if form.is_valid():
                form.instance.creator = request.user
                form.save()

                return redirect('main:instructions-list')
        else:
            form = forms.InstructionModelForm()
    else:
        raise PermissionDenied

    context = {
        'form': form,
        'url': reverse('main:add-instruction'),
        'page_heading': 'Добавяне на инструкция'
    }

    return render(request, 'main/add_update_form.html', context)


@login_required
def update_instruction(request, pk):
    instruction = get_object_or_404(models.Instruction, id=pk)

    if request.method == 'POST':
        form = forms.InstructionModelForm(
            request.POST, instance=instruction)

        if form.is_valid():
            form.save()
            return redirect('main:instructions-list')
    else:
        form = forms.InstructionModelForm(instance=instruction)

    context = {
        'form': form,
        'url': reverse('main:update-instruction', args=(pk,)),
        'page_heading': 'Редактиране на инструкция'
    }

    return render(request, 'main/add_update_form.html', context)


class InstructionDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):

    model = models.Instruction
    success_url = reverse_lazy('main:instructions-list')

    def test_func(self):
        return self.request.user.is_staff


@login_required
def instruction_xlsx(request, pk):
    if request.user.is_staff:
        instruction = get_object_or_404(models.Instruction, id=pk)

        unique_token = secrets.token_hex(32)

        xlsx_path = os.path.join(
            settings.BASE_DIR, 'main/xlsx_files/instruction.xlsx')

        unique_xlsx_path = os.path.join(
            settings.BASE_DIR, f'main/xlsx_files/instruction_{unique_token}.xlsx')

        shutil.copy(xlsx_path, unique_xlsx_path)

        company = instruction.company

        wb = load_workbook(filename=unique_xlsx_path)
        ws = wb.active

        ws['A2'] = company.name
        ws['A3'] = f'{company.city}, {company.address}'
        ws['A4'] = company.bulstat
        ws['E7'] = instruction.number
        ws['B9'] = instruction.driver.__str__()
        ws['G9'] = instruction.driver.personal_id
        ws['B11'] = instruction.car.number_plate
        ws['G18'] = instruction.driver.__str__()
        ws['G21'] = instruction.creator.__str__()
        ws['B20'] = dateformat.format(
            instruction.creation_date, formats.get_format('SHORT_DATE_FORMAT'))
        ws['A21'] = instruction.city

        ws['A25'] = company.name
        ws['A26'] = f'{company.city}, {company.address}'
        ws['A27'] = company.bulstat
        ws['E30'] = instruction.number
        ws['B32'] = instruction.driver.__str__()
        ws['G32'] = instruction.driver.personal_id
        ws['B34'] = instruction.car.number_plate
        ws['G41'] = instruction.driver.__str__()
        ws['G44'] = instruction.creator.__str__()
        ws['B43'] = dateformat.format(
            instruction.creation_date, formats.get_format('SHORT_DATE_FORMAT'))
        ws['A44'] = instruction.city

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="Instruction {instruction.number}.xlsx"'

        wb.save(response)

        os.remove(unique_xlsx_path)

        return response
    else:
        raise PermissionDenied


def letter_xlsx(course):
    company = course.company
    contractor = course.contractor

    unique_token = secrets.token_hex(32)

    xlsx_path = os.path.join(
        settings.BASE_DIR, 'main/xlsx_files/letter.xlsx')

    unique_xlsx_path = os.path.join(
        settings.BASE_DIR, f'main/xlsx_files/letter_{unique_token}.xlsx')

    shutil.copy(xlsx_path, unique_xlsx_path)

    wb = load_workbook(filename=unique_xlsx_path)
    ws = wb.active

    ws['A2'] = company.name
    ws['A3'] = f'{company.postal_code} {company.city} {company.correspondence_address}'
    ws['B4'] = company.phone_number
    ws['B5'] = company.email

    ws['G18'] = contractor.name
    ws['G19'] = contractor.correspondence_address
    ws['G20'] = f'{contractor.postal_code} {contractor.city}'
    ws['H21'] = contractor.phone_number

    ws['A26'] = company.name
    ws['A27'] = f'{company.postal_code} {company.city} {company.correspondence_address}'
    ws['B28'] = company.phone_number
    ws['B29'] = company.email

    ws['G42'] = contractor.name
    ws['G43'] = contractor.correspondence_address
    ws['G44'] = f'{contractor.postal_code} {contractor.city}'
    ws['H45'] = contractor.phone_number

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="Letter.xlsx"'

    wb.save(response)

    os.remove(unique_xlsx_path)

    return response


def receipt_xlsx(course):
    company = course.company
    contractor = course.contractor

    unique_token = secrets.token_hex(32)

    xlsx_path = os.path.join(
        settings.BASE_DIR, 'main/xlsx_files/receipt.xlsx')

    unique_xlsx_path = os.path.join(
        settings.BASE_DIR, f'main/xlsx_files/receipt_{unique_token}.xlsx')

    shutil.copy(xlsx_path, unique_xlsx_path)

    wb = load_workbook(filename=unique_xlsx_path)
    ws = wb.active

    ws['C8'] = contractor.name
    ws['E10'] = contractor.correspondence_address

    if contractor.postal_code:
        ws['C12'] = contractor.postal_code[0]
        ws['D12'] = contractor.postal_code[1]
        ws['E12'] = contractor.postal_code[2]
        ws['F12'] = contractor.postal_code[3]

    ws['J12'] = contractor.city
    ws['D15'] = course.contact_person

    ws['AC20'] = company.name
    ws['AE21'] = company.correspondence_address
    ws['AE23'] = company.province

    if company.postal_code:
        ws['AC25'] = company.postal_code[0]
        ws['AD25'] = company.postal_code[1]
        ws['AE25'] = company.postal_code[2]
        ws['AF25'] = company.postal_code[3]

    ws['AJ25'] = company.city

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="Receipt.xlsx"'

    wb.save(response)

    os.remove(unique_xlsx_path)

    return response


def official_notices_xlsx(course):
    company = course.company

    unique_token = secrets.token_hex(32)

    unique_dir_path = os.path.join(
        settings.BASE_DIR, f'main/xlsx_files/{unique_token}')

    os.mkdir(unique_dir_path)

    course_technical_inspection_xlsx_path = os.path.join(
        settings.BASE_DIR, 'main/xlsx_files/course_technical_inspection.xlsx')

    unique_course_technical_inspection_xlsx_path = os.path.join(
        unique_dir_path, 'course_technical_inspection.xlsx')

    shutil.copy(course_technical_inspection_xlsx_path,
                unique_course_technical_inspection_xlsx_path)

    wb = load_workbook(
        filename=unique_course_technical_inspection_xlsx_path)
    ws = wb.active

    ws['A1'] = company.name
    ws['A2'] = f'{company.city}, {company.address}'
    ws['A3'] = company.bulstat
    ws['E9'] = course.technical_inspection.number
    ws['B12'] = course.car.number_plate
    ws['B13'] = course.driver.__str__()
    ws['C15'] = course.technical_inspection.perpetrator.perpetrator
    ws['B18'] = dateformat.format(
        course.creation_date, formats.get_format('SHORT_DATE_FORMAT'))

    wb.save(unique_course_technical_inspection_xlsx_path)

    course_medical_examination_xlsx_path = os.path.join(
        settings.BASE_DIR, 'main/xlsx_files/course_medical_examination.xlsx')

    unique_course_medical_examination_xlsx_path = os.path.join(
        unique_dir_path, 'course_medical_examination.xlsx')

    shutil.copy(course_medical_examination_xlsx_path,
                unique_course_medical_examination_xlsx_path)

    wb = load_workbook(
        filename=unique_course_medical_examination_xlsx_path)
    ws = wb.active

    ws['A1'] = company.name
    ws['A2'] = f'{company.city}, {company.address}'
    ws['A3'] = company.bulstat
    ws['E9'] = course.medical_examination.number
    ws['A12'] = course.driver.__str__()
    ws['B13'] = course.car.number_plate
    ws['C17'] = course.medical_examination.perpetrator.perpetrator
    ws['B20'] = dateformat.format(
        course.creation_date, formats.get_format('SHORT_DATE_FORMAT'))

    wb.save(unique_course_medical_examination_xlsx_path)

    shutil.make_archive(unique_dir_path, 'zip', unique_dir_path)

    zip_path = os.path.join(
        settings.BASE_DIR, f'main/xlsx_files/{unique_token}.zip')

    response = HttpResponse(open(zip_path, 'rb'))
    response['Content-Type'] = 'application/zip'
    response['Content-Disposition'] = f'attachment; filename="Official Notices {course.medical_examination.number}.zip"'

    os.remove(zip_path)
    shutil.rmtree(unique_dir_path, ignore_errors=True)

    return response


@login_required
def course_documents_xlsx(request):
    if request.user.is_staff:
        if request.method == 'POST':
            form = forms.CourseDocumentsForm(request.POST)
            if form.is_valid():
                course = form.cleaned_data.get('course')
                document_type = form.cleaned_data.get('document_type')

                if document_type == 'Писмо':
                    response = letter_xlsx(course)
                elif document_type == 'Обратна разписка':
                    response = receipt_xlsx(course)
                elif document_type == 'Служебни бележки':
                    response = official_notices_xlsx(course)

                return response
        else:
            form = forms.CourseDocumentsForm()

        context = {
            'form': form,
            'page_heading': 'Документи за курс'
        }

        return render(request, 'main/add_update_form.html', context)
    else:
        raise PermissionDenied


@login_required
def course_date_journals_xlsx(request):
    if request.user.is_staff:
        if request.method == 'POST':
            form = forms.CourseDateJournalForm(request.POST)

            if form.is_valid():
                journal_type = form.cleaned_data['journal_type']
                from_date = form.cleaned_data['from_date']
                to_date = form.cleaned_data['to_date']
                company = form.cleaned_data['company']

                if journal_type == 'technical_inspection':
                    technical_inspections = models.CourseTechnicalInspection.objects.filter(
                        creation_date__range=[from_date, to_date]
                    )

                    unique_token = secrets.token_hex(32)

                    xlsx_path = os.path.join(
                        settings.BASE_DIR, 'main/xlsx_files/course_technical_inspection_journal.xlsx')

                    unique_xlsx_path = os.path.join(
                        settings.BASE_DIR, f'main/xlsx_files/course_technical_inspection_journal_{unique_token}.xlsx')

                    shutil.copy(xlsx_path, unique_xlsx_path)

                    heading = f'{company.name}, {company.city}, ЕИК {company.bulstat}'

                    wb = load_workbook(filename=unique_xlsx_path)
                    ws = wb.active

                    ws['A2'] = heading

                    for i in range(0, len(technical_inspections)):
                        ws[f'A{i + 6}'] = technical_inspections[i].number
                        ws[f'B{i + 6}'] = technical_inspections[i].course.car.__str__()
                        ws[f'C{i + 6}'] = dateformat.format(
                            technical_inspections[i].creation_date, formats.get_format('SHORT_DATE_FORMAT'))
                        ws[f'F{i + 6}'] = technical_inspections[i].perpetrator.perpetrator

                    response = HttpResponse(
                        content_type='application/vnd.ms-excel')
                    response['Content-Disposition'] = f'attachment; filename="Technical Inspections Journal.xlsx"'

                    wb.save(response)

                    os.remove(unique_xlsx_path)

                elif journal_type == 'medical_examination':
                    medical_examinations = models.CourseMedicalExamination.objects.filter(
                        creation_date__range=[from_date, to_date]
                    )

                    unique_token = secrets.token_hex(32)

                    xlsx_path = os.path.join(
                        settings.BASE_DIR, 'main/xlsx_files/course_medical_journal.xlsx')

                    unique_xlsx_path = os.path.join(
                        settings.BASE_DIR, f'main/xlsx_files/course_medical_journal_{unique_token}.xlsx')

                    shutil.copy(xlsx_path, unique_xlsx_path)

                    heading = f'{company.name}, {company.city}, ЕИК {company.bulstat}'

                    wb = load_workbook(filename=unique_xlsx_path)
                    ws = wb.active

                    ws['A2'] = heading

                    for i in range(0, len(medical_examinations)):
                        ws[f'A{i + 6}'] = medical_examinations[i].number
                        ws[f'B{i + 6}'] = medical_examinations[i].course.driver.__str__()
                        ws[f'C{i + 6}'] = dateformat.format(
                            medical_examinations[i].creation_date, formats.get_format('SHORT_DATE_FORMAT'))
                        ws[f'F{i + 6}'] = medical_examinations[i].perpetrator.perpetrator

                    response = HttpResponse(
                        content_type='application/vnd.ms-excel')
                    response['Content-Disposition'] = f'attachment; filename="Medical Journal.xlsx"'

                    wb.save(response)

                    os.remove(unique_xlsx_path)

                return response
        else:
            form = forms.CourseDateJournalForm()

        context = {
            'form': form,
            'page_heading': 'Сваляне на дневник'
        }

        return render(request, 'main/add_update_form.html', context)
    else:
        raise PermissionDenied


class TagAutoResponseView(AutoResponseView):
    def get(self, request, *args, **kwargs):
        self.widget = self.get_widget_or_404()
        self.term = kwargs.get('term', request.GET.get('term', ''))
        self.object_list = self.get_queryset()

        context = self.get_context_data()
        field_name = self.kwargs['field_name']

        return JsonResponse({
            'results': [
                {
                    'text': self.widget.label_from_instance(obj),
                    'id': getattr(obj, field_name),
                }
                for obj in context['object_list']
            ],
            'more': context['page_obj'].has_next()
        })
